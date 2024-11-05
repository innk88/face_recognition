import face_recognition
import cv2
import numpy as np
import os
from openpyxl import load_workbook
from datetime import datetime


base_dir = 'known_faces'
attendance_file = 'посещаемость.xlsx'  


def load_group_faces(group_name):
    group_path = os.path.join(base_dir, group_name)
    known_faces_encodings = []
    known_faces_names = []
    
    for student_name in os.listdir(group_path):
        student_path = os.path.join(group_path, student_name)
        if os.path.isdir(student_path):
            encoding_path = os.path.join(student_path, f"{student_name}_encoding.npy")
            if os.path.exists(encoding_path):
                face_encoding = np.load(encoding_path)
                known_faces_encodings.append(face_encoding)
                known_faces_names.append(student_name) 
    return known_faces_encodings, known_faces_names


def select_group():
    groups = [group for group in os.listdir(base_dir) if os.path.isdir(os.path.join(base_dir, group))]
    print("Выберите группу из списка:")
    for i, group in enumerate(groups):
        print(f"{i + 1}. {group}")
    
    while True:
        try:
            choice = int(input("Введите номер группы: ")) - 1
            if 0 <= choice < len(groups):
                return groups[choice]
            else:
                print("Неверный выбор. Попробуйте снова.")
        except ValueError:
            print("Пожалуйста, введите корректный номер.")


def mark_attendance(name, group_name):
    date_today = datetime.today().strftime('%Y-%m-%d')
    workbook = load_workbook(attendance_file)
    
    
    if group_name not in workbook.sheetnames:
        print(f"Лист '{group_name}' не найден в файле Excel.")
        return
    
    sheet = workbook[group_name]
    
    
    for row in sheet.iter_rows(min_row=2, values_only=False):
        if row[0].value == name:
            date_column = None
            for cell in sheet[1]: 
                if cell.value == date_today:
                    date_column = cell.column
                    break
            if date_column is None:
                date_column = sheet.max_column + 1
                sheet.cell(row=1, column=date_column, value=date_today)
            sheet.cell(row=row[0].row, column=date_column, value="1")
            print(f"{name} отмечен как присутствующий.")
            break
    workbook.save(attendance_file)
    workbook.close()


group_name = select_group()
print(f"Вы выбрали группу: {group_name}")

known_faces_encodings, known_faces_names = load_group_faces(group_name)

input_image_path = 'class_photo.jpg'
input_image = cv2.imread(input_image_path)

face_locations = face_recognition.face_locations(input_image)
face_encodings = face_recognition.face_encodings(input_image, face_locations)

for (top, right, bottom, left), face_encoding in zip(face_locations, face_encodings):
    matches = face_recognition.compare_faces(known_faces_encodings, face_encoding, tolerance=0.4)
    name = "Unknown"
    if True in matches:
        first_match_index = matches.index(True)
        name = known_faces_names[first_match_index]
        mark_attendance(name, group_name) 
    cv2.rectangle(input_image, (left, top), (right, bottom), (0, 255, 0), 2)
    cv2.putText(input_image, name, (left, top - 10), cv2.FONT_HERSHEY_SIMPLEX, 0.75, (0, 255, 0), 2)

cv2.imshow('Recognition', input_image)
cv2.waitKey(0)
cv2.destroyAllWindows()
